import tkinter as tk
from tkinter import filedialog
import openpyxl
from selenium import webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import time
from selenium.webdriver.chrome.options import Options
from datetime import datetime, timedelta

#输入Excel文件
root=tk.Tk()
root.withdraw()
file_path=filedialog.askopenfilename(title="选择要处理的Excel文件",filetypes=[("Excel文件","*.xlsx;*.xls")])

#数据紧急留存
try_phone=None
global_code_value=None
yes=None

flag_1=0
flag_2=0
# 获取账号信息
with pd.ExcelFile(file_path) as xls:
    df = pd.read_excel(xls)
# 遍历每个账号
    for index, row in df.iterrows():
        # 获取对应列的名称
        tel_col_name = df.columns[df.iloc[0] == '电话'].values[0]
        code_col_name = df.columns[df.iloc[0] == '统一社会信用代码'].values[0]
        tel_more_name = df.columns[df.iloc[0] == '更多电话'].values[0]
        person_name = df.columns[df.iloc[0] == '法定代表人'].values[0]
        # 跳过第一行、为空的和无数据的
        if pd.isna(row[tel_col_name]) or row[tel_col_name] == '' or index==0:
            continue
        # 获取电话和统一代码的值
        tel_value = row[tel_col_name]
        code_value = row[code_col_name]
        tel_more_value=None
        if pd.isna(row[tel_more_name])!=1 or row[tel_more_name]!='':
            tel_more_value = row[tel_more_name]
        person_value=row[person_name]
        #拼接数据
        if pd.isna(row[tel_more_name])!=1 and tel_more_value != '-':
            tel_value = f"{tel_value}；{tel_more_value}"
        if pd.isna(row[person_name])!=1 and person_value != '-':
            tel_value = f"{tel_value}；{person_value}"
        # print(tel_value)
        #留存数据赋值
        global_code_value=code_value

        #打开Chrome驱动（需提前安装驱动）
        options = Options()
        options.add_argument('--incognito')  # 添加无痕模式启动参数
        driver = webdriver.Chrome(options=options)
        driver.implicitly_wait(3)  # 设置隐式等待时间为10秒
        wait = WebDriverWait(driver, 10)
        driver.get("https://tpass.guangdong.chinatax.gov.cn:8443/#/login?redirect_uri=https%3A%2F%2Fetax.guangdong.chinatax.gov."
               "cn%2Fsso%2Flogin%3Fservice%3Dhttps%253A%252F%252Fetax.guangdong.chinatax.gov.cn%252Fxxmh%252Fhtml%252Findex_l"
               "ogin.html%26v%3D2&client_id=f91863cc09c75f6b881b8f9953035b6f&response_type=code&state=test")
        time.sleep(5)
        #跳转密码登录
        # wait.until(EC.presence_of_element_located((By.XPATH, '//div[@class="scanTxt"]')))
        # driver.find_element(by=By.XPATH, value='//div[@class="scanTxt"]').click()
        wait.until(EC.presence_of_element_located((By.XPATH,'//input[@placeholder="统一社会信用代码/纳税人识别号"]')))
        input_element = driver.find_element(by=By.XPATH, value='//input[@placeholder="统一社会信用代码/纳税人识别号"]')
        username = driver.find_element(by=By.XPATH, value='//input[@placeholder="居民身份证号码/手机号码/用户名"]')
        password = driver.find_element(by=By.XPATH, value='//input[@placeholder="个人用户密码"]')
        #测试数据
        # code_value='9144060'
        # tel_value='138'
        #输入信用代码和密码
        input_element.send_keys(str(code_value))
        password.send_keys('Aa123456')


        #输入账号（先电话再名称）
        for part in tel_value.split('；'):
            if str(part).startswith('0') or str(part).startswith('4'):
                continue
            if (str(part).isdigit())!=1:
                flag_2=1
            else:flag_2=0
            try:
                username.clear()
                username.send_keys(str(part))
                try_phone=str(part)
                #滑动验证 !!!!!!!
                # try:
                #     driver.find_element(by=By.XPATH, value='//div[@class="handler animate"]')
                # except NoSuchElementException:
                #     continue
                wait.until(EC.presence_of_element_located((By.XPATH, '//div[@class="handler animate"]')))
                slide = driver.find_element(by=By.XPATH, value='//div[@class="handler animate"]')
                ActionChains(driver).drag_and_drop_by_offset(slide, 366, 0).perform()
                driver.find_element(by=By.XPATH, value='//button[@class="el-button loginCls el-button--primary"]').click()
                #检测是否登陆成功
                try:
                    driver.find_element(by=By.XPATH, value='//span[contains(text(),"确认")]').click()
                    yes = 1
                    break
                except Exception as x:
                    pass

                try:
                    driver.find_element(by=By.ID, value='wdxx')
                    yes = 1
                    break
                except NoSuchElementException:
                    pass
                yes = 0
            except Exception as e:
                yes = 0
                break

        try:
            # 删掉已经测过的（至少等当前账号操作结束再关闭进程）
            # df.drop(df[df[tel_col_name] == row[tel_col_name]].index, inplace=True)
            # df.drop(df[df[code_col_name] == row[code_col_name]].index, inplace=True)
            # df.drop(df[df[tel_more_name] == row[tel_more_name]].index, inplace=True)
            # df.drop(df[df[person_name] == row[person_name]].index, inplace=True)
            df.replace(row[tel_col_name], '', inplace=True)
            df.replace(row[code_col_name], '', inplace=True)
            df.replace(row[tel_more_name], '', inplace=True)
            if flag_1 == 0:
                df.replace(row[person_name], '标志', inplace=True)
                flag_1 += 1
            else:
                df.replace(row[person_name], '', inplace=True)

            # 将处理后的数据写回 Excel 表格
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

        # except Exception as e:
        #     print(f"写回失败：{e}")
            #如果没登录进去，下一个
            if(yes==0):
                continue
            #如果能进去，检测是否为新用户
            try:
                elements_2 = driver.find_elements(By.XPATH,'//span[text()="《个人信息保护告知同意书》")]')
                driver.find_element(by=By.XPATH,value='//span[text()="确认"]').click()
            except Exception:
                pass
            # 继续爬数据
            wait.until(EC.presence_of_element_located((By.LINK_TEXT, '银税互动平台')))
            # 跳转页面
            driver.find_element(by=By.LINK_TEXT, value='银税互动平台').click()
            # 跳转参照域
            handles = driver.window_handles
            driver.switch_to.window(handles[-1])

            # 先爬第二个
            wait.until(EC.presence_of_element_located((By.ID, 'cx')))
            driver.find_element(by=By.ID, value='cx').click()
            time.sleep(2)
            # 跳转入内嵌页面
            driver.switch_to.frame("iframe")
            # 通过js修改时间输入框参数
            js = """
                                var date = document.getElementById("ssrq_q");
                                date.readOnly = false;
                                date.value = arguments[0];
                             """
            driver.execute_script(js, "2020-01-01")
            driver.find_element(by=By.XPATH, value='//input[@value="查 询"]').click()
            wait.until(EC.presence_of_element_located((By.XPATH, '//table[@id="userList"]')))
            # 定位到数据
            table = driver.find_element(by=By.XPATH, value='//table[@id="userList"]')
            tbody = driver.find_element(by=By.XPATH, value='//table[@id="userList"]/tbody')
            data = []
            # 爬取并存储数据
            for tr in tbody.find_elements(By.XPATH, './tr'):
                row = []
                tds = tr.find_elements(By.XPATH, './td')
                for td in tds[2:]:
                    row.append(td.text)
                data.append(row)

            # 再爬第一个
            driver.switch_to.default_content()
            wait.until(EC.presence_of_element_located((By.ID, 'sq')))
            driver.find_element(by=By.ID, value='sq').click()
            driver.switch_to.frame("iframe")
            wait.until(EC.presence_of_element_located((By.ID, 'checksqs')))
            driver.find_element(by=By.ID, value='checksqs').click()
            time.sleep(1)
            driver.find_element(by=By.XPATH, value='//p[text()="下一步"]').click()
            time.sleep(2)
            sb_frame = driver.find_element(by=By.XPATH, value='//iframe[@scrolling="auto"]')
            driver.switch_to.frame(sb_frame)
            driver.switch_to.frame("iframe")
            driver.find_element(by=By.XPATH, value='//img[@name="00000032"]').click()
            wait.until(EC.presence_of_element_located((By.XPATH, '//img[@name="00000004"]')))
            driver.find_element(by=By.XPATH, value='//img[@name="00000004"]').click()
            time.sleep(1)
            driver.switch_to.parent_frame()
            wait.until(EC.presence_of_element_located((By.XPATH, '//span[@class="qd_btn"]')))
            driver.find_element(by=By.XPATH, value='//span[@class="qd_btn"]').click()
            driver.switch_to.parent_frame()
            wait.until(EC.presence_of_element_located((By.XPATH, '//span[@id="yulan"]')))
            driver.find_element(by=By.XPATH, value='//span[@id="yulan"]').click()
            wait.until(EC.presence_of_element_located((By.ID, 'nsxxid')))
            driver.find_element(by=By.ID, value='nsxxid').click()
            # 寻找数据并进行处理
            # 纳税
            wait.until(EC.presence_of_element_located((By.XPATH, '//table[@id="tb9"]/tbody/tr[1]')))
            tr_location = driver.find_element(by=By.XPATH, value='//table[@id="tb9"]/tbody')
            tr_num = tr_location.find_elements(By.TAG_NAME, 'tr')
            # 第一套
            # tax_2023 = 0
            # tax_2022 = 0
            # tax_2021 = 0
            # tax_2020 = 0
            # 第二套
            recent_year = 0
            fore_recent_year = 0
            # 实际数据
            tax = []
            for i in range(len(tr_num)):
                td_text1 = driver.find_element(by=By.XPATH, value=f"//table[@id='tb9']/tbody/tr[{i + 1}]/td[2]").text
                td_text2 = driver.find_element(by=By.XPATH, value=f"//table[@id='tb9']/tbody/tr[{i + 1}]/td[6]").text
                td_text3 = driver.find_element(by=By.XPATH, value=f"//table[@id='tb9']/tbody/tr[{i + 1}]/td[8]").text
                td_text4 = driver.find_element(by=By.XPATH, value=f"//table[@id='tb9']/tbody/tr[{i + 1}]/td[5]").text
                if td_text4.startswith('被'):
                    continue
                else:
                    # 第一套，按年份
                    # if td_text1.startswith('2023'):
                    #     if td_text2.startswith('增值'):
                    #         tax_2023 += float(td_text3)
                    # if td_text1.startswith('2022'):
                    #     if td_text2.startswith('增值'):
                    #         tax_2022 += float(td_text3)
                    # if td_text1.startswith('2021'):
                    #     if td_text2.startswith('增值'):
                    #         tax_2021 += float(td_text3)
                    # if td_text1.startswith('2020'):
                    #     if td_text2.startswith('增值'):
                    #         tax_2020 += float(td_text3)
                    # 第二套，按近一年
                    compare_date = datetime.strptime(td_text1, '%Y-%m-%d')
                    now = datetime.now()
                    one_year_ago = now - timedelta(days=365)
                    two_year_ago = one_year_ago - timedelta(days=365)
                    if one_year_ago <= compare_date <= now:
                        if td_text2.startswith('增值'):
                            recent_year += float(td_text3)
                    elif two_year_ago <= compare_date <= one_year_ago:
                        if td_text2.startswith('增值'):
                            fore_recent_year += float(td_text3)
            # 第一套
            # tax.append(tax_2023)
            # tax.append(tax_2022)
            # tax.append(tax_2021)
            # tax.append(tax_2020)
            # 第二套
            tax.append(recent_year)
            tax.append(fore_recent_year)
            invoice = []
            try:
                # 开票
                wait.until(EC.presence_of_element_located((By.XPATH, '//table[@id="tb32"]/tbody/tr[1]')))
                tr_location = driver.find_element(by=By.XPATH, value='//table[@id="tb32"]/tbody')
                tr_num = tr_location.find_elements(By.TAG_NAME, 'tr')
                # 第一套
                # invoice_2023 = 0
                # invoice_2022 = 0
                # invoice_2021 = 0
                # invoice_2020 = 0
                # 第二套
                recent_year2 = 0
                fore_recent_year2 = 0
                for i in range(len(tr_num)):
                    td_text1 = driver.find_element(by=By.XPATH, value=f"//table[@id='tb32']/tbody/tr[{i + 1}]/td[2]").text
                    td_text2 = driver.find_element(by=By.XPATH, value=f"//table[@id='tb32']/tbody/tr[{i + 1}]/td[5]").text
                    td_text3 = driver.find_element(by=By.XPATH, value=f"//table[@id='tb32']/tbody/tr[{i + 1}]/td[7]").text
                    # 第一套，按年份
                    # if td_text1.startswith('2023'):
                    #     if td_text2.startswith('增值'):
                    #         invoice_2023 += float(td_text3)
                    # if td_text1.startswith('2022'):
                    #     if td_text2.startswith('增值'):
                    #         invoice_2022 += float(td_text3)
                    # if td_text1.startswith('2021'):
                    #     if td_text2.startswith('增值'):
                    #         invoice_2021 += float(td_text3)
                    # if td_text1.startswith('2020'):
                    #     if td_text2.startswith('增值'):
                    #         invoice_2020 += float(td_text3)
                    # 第二套，按近一年
                    compare_date = datetime.strptime(td_text1, '%Y-%m-%d')
                    now = datetime.now()
                    one_year_ago = now - timedelta(days=365)
                    two_year_ago = one_year_ago - timedelta(days=365)
                    if one_year_ago <= compare_date <= now:
                        if td_text2.startswith('增值'):
                            recent_year2 += float(td_text3)
                    elif two_year_ago <= compare_date <= one_year_ago:
                        if td_text2.startswith('增值'):
                            fore_recent_year2 += float(td_text3)

                # 第一套
                # invoice.append(invoice_2023)
                # invoice.append(invoice_2022)
                # invoice.append(invoice_2021)
                # invoice.append(invoice_2020)
                # 第二套
                invoice.append(recent_year2)
                invoice.append(fore_recent_year2)
            except Exception:
                invoice=[0,0]
            # 信用评级
            wait.until(EC.presence_of_element_located((By.ID, 'xypjxxid')))
            driver.find_element(by=By.ID, value='xypjxxid').click()
            wait.until(EC.presence_of_element_located((By.XPATH, '//table[@id="tb31"]/tbody/tr[1]')))
            tr_location = driver.find_element(by=By.XPATH, value='//table[@id="tb31"]/tbody')
            tr_num = tr_location.find_elements(By.TAG_NAME, 'tr')
            xypj_2019 = "2019 "
            xypj_2022 = "2022 "
            xypj_2021 = "2021 "
            xypj_2020 = "2020 "
            xypj = []
            for i in range(len(tr_num)):
                td_text1 = driver.find_element(by=By.XPATH, value=f"//table[@id='tb31']/tbody/tr[{i + 1}]/td[4]").text
                td_text3 = driver.find_element(by=By.XPATH, value=f"//table[@id='tb31']/tbody/tr[{i + 1}]/td[5]").text
                if td_text1.startswith('2022'):
                    xypj_2022 += str(td_text3)
                if td_text1.startswith('2021'):
                    xypj_2021 += str(td_text3)
                if td_text1.startswith('2020'):
                    xypj_2020 += str(td_text3)
                if td_text1.startswith('2019'):
                    xypj_2019 += str(td_text3)
            xypj.append(xypj_2022)
            xypj.append(xypj_2021)
            xypj.append(xypj_2020)
            xypj.append(xypj_2019)
            driver.switch_to.window(driver.window_handles[-2])
            bank_info = []
            company_info=[]
            wait.until(EC.presence_of_element_located((By.ID, 'wycx')))
            driver.find_element(by=By.ID, value='wycx').click()
            wait.until(EC.presence_of_element_located((By.XPATH, '//h4[text()="一户式查询"]/parent::a')))
            driver.find_element(by=By.XPATH, value='//h4[text()="一户式查询"]/parent::a').click()
            driver.switch_to.frame("ifrMain")
            driver.switch_to.frame("cxtable")
            driver.switch_to.frame("nsrxx")
            try:
                wait.until(
                    EC.presence_of_element_located((By.XPATH, '//td[text()="纳税人名称："]/following-sibling::td[1]')))
                company = driver.find_element(by=By.XPATH, value='//td[text()="纳税人名称："]/following-sibling::td[1]')
                company_info.append(company.text)
            except Exception as whats:
                print(whats)
                company_info.append('无')
            driver.switch_to.default_content()

            driver.switch_to.frame("ifrMain")
            wait.until(EC.presence_of_element_located((By.XPATH, '//a[text()="存款账户账号报告查询"]')))
            driver.find_element(by=By.XPATH, value='//a[text()="存款账户账号报告查询"]').click()
            # 跳转入内嵌页面
            driver.switch_to.frame("cxtable")
            try:
                wait.until(
                    EC.presence_of_element_located((By.XPATH, '//td[text()="基本存款账户"]/following-sibling::td[1]')))
                bank = driver.find_element(by=By.XPATH, value='//td[text()="基本存款账户"]/following-sibling::td[1]')
                bank_info.append(bank.text)
            except Exception as what:
                print(what)
                bank_info.append('无')
            driver.switch_to.default_content()
            driver.find_element(by=By.XPATH, value='//a[text()="返回主页"]').click()
            try:
                # 爬取数据
                wait.until(EC.presence_of_element_located((By.ID, 'wdxx')))
                test = driver.find_element(by=By.ID, value='wdxx')
                test.click()
                wait.until(EC.presence_of_element_located((By.XPATH, '//h4[text()="账户中心"]/parent::a')))
                driver.find_element(by=By.XPATH, value='//h4[text()="账户中心"]/parent::a').click()
                # 如果为新用户，进行该处理(加工中。。。不行)
                try:
                    wait.until(EC.presence_of_element_located((By.XPATH, '//span[text()="用户名统一设置"]')))
                    driver.find_element(by=By.XPATH, value='//span[text()="用户名统一设置"]')
                    driver.find_element(by=By.XPATH, value='//span[text()="确认"]').click()
                    wait.until(EC.presence_of_element_located((By.XPATH, '//input[[@placeholder="请设置密码"]]')))
                    driver.find_element(by=By.XPATH, value='//input[[@placeholder="请设置密码"]]').send_keys('Aa123456')
                    driver.find_element(by=By.XPATH, value='//input[[@placeholder="请确认密码"]]').send_keys('Aa123456')
                    driver.find_element(by=By.XPATH, value='//span[text()="确认"]').click()
                    time.sleep(2)
                    driver.quit()
                    driver.get(
                        "https://etax.guangdong.chinatax.gov.cn/sso/login?service=https%3A%2F%2Fetax.guangdong.chinatax.gov."
                        "cn%2Fsbzx-cjpt-web%2Fsb%2Fhtml.do%3Fredirect_uri%3DcxstysbYdy%26zrrBz%3DY")
                    time.sleep(5)
                    # 跳转密码登录
                    driver.find_element(by=By.XPATH, value='//div[@class="scanTxt"]').click()
                    wait.until(
                        EC.presence_of_element_located((By.XPATH, '//input[@placeholder="统一社会信用代码/纳税人识别号"]')))
                    input_element = driver.find_element(by=By.XPATH,
                                                        value='//input[@placeholder="统一社会信用代码/纳税人识别号"]')
                    username = driver.find_element(by=By.XPATH, value='//input[@placeholder="居民身份证号码/手机号码/用户名"]')
                    password = driver.find_element(by=By.XPATH, value='//input[@placeholder="个人用户密码"]')
                    input_element.send_keys(global_code_value)
                    username.send_keys(try_phone)
                    password.send_keys('Aa123456')
                    # 滑动验证
                    wait.until(EC.presence_of_element_located((By.XPATH, '//div[@class="handler animate"]')))
                    slide = driver.find_element(by=By.XPATH, value='//div[@class="handler animate"]')
                    ActionChains(driver).drag_and_drop_by_offset(slide, 520, 0).perform()
                    driver.find_element(by=By.XPATH, value='//button[@class="el-button loginCls el-button--primary"]').click()
                    time.sleep(2)
                    # 检测是否登陆成功
                    try:
                        driver.find_element(by=By.XPATH, value='//span[contains(text(),"确认")]').click()
                    except NoSuchElementException:
                        pass

                    try:
                        driver.find_element(by=By.ID, value='wdxx').click()
                    except NoSuchElementException:
                        pass
                    wait.until(EC.presence_of_element_located((By.XPATH, '//h4[text()="账户中心"]/parent::a')))
                    driver.find_element(by=By.XPATH, value='//h4[text()="账户中心"]/parent::a').click()
                except Exception as c:
                    pass

                # 继续爬取数据
                basic = []
                wait.until(EC.presence_of_element_located((By.XPATH, '//img[@class="header-eye"]')))
                driver.find_element(by=By.XPATH, value='//img[@class="header-eye"]').click()
                time.sleep(3)
                # company_name = driver.find_element(by=By.XPATH,
                #                                    value='//div[text()="纳税人名称"]/following-sibling::div//span[1]')
                credit_code = driver.find_element(by=By.XPATH,
                                                  value='//div[text()="统一社会信用代码"]/following-sibling::div//span[1]')
                Location = driver.find_element(by=By.XPATH, value='//div[text()="生产经营地址"]/following-sibling::div[1]')
                Law_person = driver.find_element(by=By.XPATH,
                                                 value='//div[text()="法定代表人姓名"]/following-sibling::div//span[1]')
                Law_person_phone = driver.find_element(by=By.XPATH,
                                                       value='//div[text()="法定代表人手机号"]/following-sibling::div//span[1]')
                Money_person = driver.find_element(by=By.XPATH,
                                                   value='//div[text()="财务负责人姓名"]/following-sibling::div//span[1]')
                Money_person_phone = driver.find_element(by=By.XPATH,
                                                         value='//div[text()="财务负责人手机号"]/following-sibling::div//span[1]')
                # basic.append(company_name.text)
                basic.append(credit_code.text)
                basic.append(Location.text)
                basic.append(Law_person.text)
                basic.append(Law_person_phone.text)
                basic.append(Money_person.text)
                basic.append(Money_person_phone.text)

                wait.until(EC.presence_of_element_located((By.XPATH, '//span[text()="个人信息管理"]')))
                driver.find_element(by=By.XPATH, value='//span[text()="个人信息管理"]').click()
                wait.until(EC.presence_of_element_located((By.XPATH, '//div[text()="姓名"]/following-sibling::div//span[1]')))
                driver.find_element(by=By.XPATH, value='//img[@class="header-eye"]').click()
                time.sleep(3)
                person_name = driver.find_element(by=By.XPATH, value='//div[text()="姓名"]/following-sibling::div//span[1]')
                phone = driver.find_element(by=By.XPATH, value='//div[text()="手机号码"]/following-sibling::div//span[1]')
                ID_card = driver.find_element(by=By.XPATH, value='//div[text()="证件号码"]/following-sibling::div//span[1]')
                basic.append(person_name.text)
                basic.append(phone.text)
                basic.append(ID_card.text)

                driver.find_element(by=By.XPATH, value='//span[text()="返回"]').click()



                # 第一次写入Excel
                file_paths = '.\\data.xlsx'
                #如果文件存在，打开，否则新建
                try:
                    wb = openpyxl.load_workbook(file_paths)
                except FileNotFoundError:
                    wb = openpyxl.Workbook()
                sheet = wb.active
                column = sheet['A']
                # 获取起始行
                start_row = 1
                for cell in column:
                    if cell.value is None and cell.offset(row=1, column=0).value is None:
                        break
                    start_row += 1
                # print(start_row)
                # print('分割线')
                # 写入(可修改)
                ws = wb.active
                j=0
                for i in range(len(basic)):
                    if j>=6:
                        ws.cell(row=2 + start_row, column=j -5, value=basic[i])
                    else:
                        ws.cell(row=1 + start_row, column=j + 1, value=basic[i])
                    j+=1
                extre1=j-6
                extre2=j-5
                ws.cell(row=2+start_row,column=j-5,value=bank_info[0])
                ws.cell(row=2+start_row,column=j-4,value=company_info[0])

                # 保存
                wb.save(file_paths)
            except Exception as why:
                print(why)
                # 第一次写入Excel
                file_paths = '.\\data.xlsx'
                # 如果文件存在，打开，否则新建
                try:
                    wb = openpyxl.load_workbook(file_paths)
                except FileNotFoundError:
                    wb = openpyxl.Workbook()
                sheet = wb.active
                column = sheet['A']
                # 获取起始行
                start_row = 1
                for cell in column:
                    if cell.value is None and cell.offset(row=1, column=0).value is None:
                        break
                    start_row += 1
                # print(start_row)
                # print('分割线')
                # 写入(可修改)
                ws = wb.active
                ws.cell(row=1 + start_row, column=1, value=global_code_value)
                ws.cell(row=1 + start_row, column=2, value=try_phone)
                ws.cell(row=1 + start_row, column=3, value=bank_info[0])
                ws.cell(row=1 + start_row, column=4, value=company_info[0])

                # 保存
                wb.save(file_paths)

            # 第二次写入Excel
            file_paths = '.\\data.xlsx'
            # 如果文件存在，打开，否则新建
            try:
                wb = openpyxl.load_workbook(file_paths)
            except FileNotFoundError:
                wb = openpyxl.Workbook()
            flag='$$$'
            sheet = wb.active
            column = sheet['A']
            # 获取起始行
            start_row = 1
            for cell in column:
                if cell.value is None and cell.offset(row=1, column=0).value is None:
                    break
                start_row += 1
            # print(start_row)
            # 写入(可修改)
            ws = wb.active
            #第一个
            for i in range(4):
                if i==0:
                    for j in range(2):
                        ws.cell(row=j + start_row, column=i + 1, value=tax[j])
                if i==1:
                    for j in range(2):
                        ws.cell(row=j + start_row, column=i + 1, value=invoice[j])
                if i==2:
                    for j in range(2):
                        ws.cell(row=j + start_row, column=i + 1, value=xypj[j])
                if i==3:
                    for j in range(2):
                        ws.cell(row=j + start_row, column=i + 1, value=xypj[j+2])
            ws.cell(row=1 + start_row, column=4 + 1, value=flag)
            if(flag_2==1):
                ws.cell(row=1 + start_row, column=5 + 1, value='&&&')
            #第二个
            for i in range(len(data)):
                for j in range(len(data[i])):
                    ws.cell(row=i+2+start_row,column=j+1,value=data[i][j])

            # 保存
            wb.save(file_paths)
            # 获取账号信息
            df = pd.read_excel(file_path, engine='openpyxl')

            driver.quit()
        except Exception as i:
            print(i)
            #紧急保存
            wrong=[]
            wrong.append(global_code_value)
            wrong.append(try_phone)
            # 写入Excel，有则打开，无则新建
            try:
                wb = openpyxl.load_workbook('.\\wrong.xlsx')
            except FileNotFoundError:
                wb = openpyxl.Workbook()
            sheet = wb.active
            column = sheet['A']
            # 获取起始行
            start_row = 1
            for cell in column:
                if cell.value is None and cell.offset(row=1, column=0).value is None:
                    break
                start_row += 1
            # 写入(可修改)
            ws = wb.active
            for j in range(2):
                ws.cell(row=1 + start_row, column=j + 1, value=wrong[j])
            # 保存
            wb.save('.\\wrong.xlsx')
            driver.quit()
            # 获取账号信息
            df = pd.read_excel(file_path, engine='openpyxl')
            continue